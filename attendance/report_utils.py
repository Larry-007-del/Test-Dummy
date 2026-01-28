from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime


def generate_attendance_pdf(attendances, course=None, date_range=None):
    """Generate PDF report for attendance records"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1976d2'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = f"Attendance Report"
    if course:
        title += f" - {course.name}"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Metadata
    meta_style = styles['Normal']
    if date_range:
        elements.append(Paragraph(f"Period: {date_range}", meta_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", meta_style))
    elements.append(Paragraph(f"Total Records: {len(attendances)}", meta_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [['#', 'Student', 'Student ID', 'Course', 'Lecturer', 'Date', 'Check-in Time']]
    
    for idx, attendance in enumerate(attendances, 1):
        student = attendance.student
        course_obj = attendance.course
        lecturer = course_obj.lecturer if course_obj else None
        
        data.append([
            str(idx),
            student.name if student else 'N/A',
            student.student_id if student else 'N/A',
            course_obj.name if course_obj else 'N/A',
            lecturer.name if lecturer else 'N/A',
            attendance.date.strftime('%Y-%m-%d') if attendance.date else 'N/A',
            attendance.date.strftime('%I:%M %p') if attendance.date else 'N/A',
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_attendance_excel(attendances, course=None, date_range=None):
    """Generate Excel report for attendance records"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    headers = ['#', 'Student Name', 'Student ID', 'Course', 'Course Code', 'Lecturer', 'Date', 'Check-in Time']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for idx, attendance in enumerate(attendances, 1):
        student = attendance.student
        course_obj = attendance.course
        lecturer = course_obj.lecturer if course_obj else None
        
        row = [
            idx,
            student.name if student else 'N/A',
            student.student_id if student else 'N/A',
            course_obj.name if course_obj else 'N/A',
            course_obj.course_code if course_obj else 'N/A',
            lecturer.name if lecturer else 'N/A',
            attendance.date.strftime('%Y-%m-%d') if attendance.date else 'N/A',
            attendance.date.strftime('%I:%M %p') if attendance.date else 'N/A',
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
